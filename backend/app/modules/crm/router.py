"""
CRM API endpoints: Customer CRUD, import, interactions, stats, duplicates.
"""
from typing import List
from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import io
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from app.core.database import get_db
from app.core.security import get_current_user
from app.core.deps import PaginationParams, CustomerFilterParams
from app.modules.crm.service import CRMService
from app.modules.crm.import_service import import_from_file, import_vehicles_from_file
from app.modules.crm.schemas import (
    CustomerCreate, CustomerUpdate, CustomerResponse,
    CustomerListResponse, InteractionCreate, InteractionResponse,
    ImportResult, CRMStats, DuplicateGroup,
    ContactCreate, ContactUpdate, ContactResponse,
    ProformaCreate, ProformaUpdate, ProformaResponse,
    VehicleResponse,
)

router = APIRouter(prefix="/api/crm", tags=["CRM"])


@router.get("/customers", response_model=CustomerListResponse)
def list_customers(
    pagination: PaginationParams = Depends(),
    filters: CustomerFilterParams = Depends(),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Müşteri listesi — filtreleme ve sayfalama destekli."""
    return CRMService(db).get_customers(pagination, filters)


@router.get("/customers/map-markers", response_model=List[dict])
def get_map_markers(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Harita üzerinde gösterilecek tüm müşteri koordinatlarını hafif biçimde döner."""
    return CRMService(db).get_map_markers()


@router.get("/customers/{customer_id}", response_model=CustomerResponse)
def get_customer(customer_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Müşteri detayı."""
    return CRMService(db).get_customer(customer_id)


@router.post("/customers", response_model=CustomerResponse)
def create_customer(data: CustomerCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Yeni müşteri ekle."""
    return CRMService(db).create_customer(data)


@router.put("/customers/{customer_id}", response_model=CustomerResponse)
def update_customer(customer_id: int, data: CustomerUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Müşteri güncelle."""
    return CRMService(db).update_customer(customer_id, data)


@router.delete("/customers/{customer_id}")
def delete_customer(customer_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Müşteri sil (kalıcı)."""
    CRMService(db).delete_customer(customer_id)
    return {"message": "Müşteri silindi"}


@router.post("/customers/bulk-delete")
def bulk_delete(data: dict, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Toplu müşteri silme."""
    ids = data.get("ids", [])
    if not ids:
        raise HTTPException(status_code=400, detail="Silinecek müşteri ID'leri gerekli")
    count = CRMService(db).bulk_delete_customers(ids)
    return {"message": f"{count} müşteri silindi", "deleted": count}


@router.delete("/customers/source/{source}")
def delete_by_source(source: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Belirli kaynaktan gelen tüm müşterileri sil."""
    count = CRMService(db).delete_by_source(source)
    return {"message": f"{count} müşteri silindi (kaynak: {source})", "deleted": count}


@router.post("/customers/import", response_model=ImportResult)
async def import_customers(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Excel/CSV dosyasından müşteri aktar."""
    return await import_from_file(file, db, current_user.id)


@router.post("/customers/check-duplicate")
def check_duplicate(data: CustomerCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Duplicate kontrolü yap."""
    return CRMService(db).check_duplicate(data)


@router.get("/customers/{customer_id}/interactions", response_model=List[InteractionResponse])
def list_interactions(customer_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Müşteri etkileşim geçmişi."""
    return CRMService(db).get_interactions(customer_id)


@router.post("/customers/{customer_id}/interactions", response_model=InteractionResponse)
def create_interaction(customer_id: int, data: InteractionCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Yeni etkileşim ekle."""
    return CRMService(db).create_interaction(customer_id, data, current_user.id)


@router.get("/stats", response_model=CRMStats)
def get_stats(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """CRM istatistikleri."""
    return CRMService(db).get_stats()


@router.get("/nearby")
def get_nearby_customers(
    lat: float,
    lon: float,
    radius: float = 5000,
    segment: str = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Konum bazlı yakındaki müşterileri getirir (saha satışı için)."""
    return CRMService(db).get_nearby_customers(lat, lon, radius, segment)


@router.get("/route-search")
def route_search(
    start_lat: float,
    start_lon: float,
    end_lat: float,
    end_lon: float,
    threshold: float = 2000,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Güzergah boyunca yakınlıktaki müşterileri listeler."""
    return CRMService(db).get_route_along_customers(start_lat, start_lon, end_lat, end_lon, threshold)




@router.get("/duplicates", response_model=List[DuplicateGroup])
def find_duplicates(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Potansiyel duplicate müşterileri bul."""
    return CRMService(db).find_all_duplicates()


# ── Contact Endpoints ──────────────────────────────────────────────────

@router.get("/customers/{customer_id}/contacts", response_model=List[ContactResponse])
def list_contacts(customer_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Firma irtibat kişileri listesi."""
    return CRMService(db).get_contacts(customer_id)


@router.post("/customers/{customer_id}/contacts", response_model=ContactResponse)
def add_contact(customer_id: int, data: ContactCreate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Firmaya yeni irtibat kişisi ekle."""
    return CRMService(db).add_contact(customer_id, data)


@router.put("/contacts/{contact_id}", response_model=ContactResponse)
def update_contact(contact_id: int, data: ContactUpdate, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """İrtibat kişisi güncelle."""
    return CRMService(db).update_contact(contact_id, data)


@router.delete("/contacts/{contact_id}")
def delete_contact(contact_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """İrtibat kişisi sil."""
    CRMService(db).delete_contact(contact_id)
    return {"message": "İrtibat kişisi silindi"}


@router.get("/contact-suggestions")
def get_contact_suggestions(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Rehberden bulunan firma-kişi gruplarını getir."""
    import json, os
    report_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(__file__)))), 'contact_groups.json')
    if not os.path.exists(report_path):
        return []
    with open(report_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data.get('groups', [])


@router.post("/customers/merge")
def merge_customers(data: dict, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Birden fazla müşteri kaydını birleştir.
    primary_id: Ana firma kaydı (kalacak olan)
    secondary_ids: Birleştirilecek kayıtlar (kişi olarak taşınıp silinecek)
    """
    primary_id = data.get("primary_id")
    secondary_ids = data.get("secondary_ids", [])
    if not primary_id or not secondary_ids:
        raise HTTPException(status_code=400, detail="primary_id ve secondary_ids gerekli")
    result = CRMService(db).merge_customers(primary_id, secondary_ids)
    return result


# ── Proforma Invoice Endpoints ──────────────────────────────────────────

@router.get("/customers/{customer_id}/proformas", response_model=List[ProformaResponse])
def list_customer_proformas(customer_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Müşteriye ait tüm proformaları listeler."""
    return CRMService(db).list_customer_proformas(customer_id)


@router.post("/customers/{customer_id}/proformas", response_model=ProformaResponse)
def create_proforma(
    customer_id: int,
    data: ProformaCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Müşteriye yeni bir proforma fatura oluşturur."""
    return CRMService(db).create_proforma(customer_id, data, current_user.id)


@router.get("/proformas/{proforma_id}", response_model=ProformaResponse)
def get_proforma(proforma_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """ID'ye göre proforma detayını getirir."""
    return CRMService(db).get_proforma(proforma_id)


@router.get("/proformas/{proforma_id}/export-excel")
def export_proforma_excel(proforma_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Proforma faturayı Excel formatında indirir."""
    service = CRMService(db)
    proforma = service.get_proforma(proforma_id)
    customer = service.get_customer(proforma.customer_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PROFORMA TEKLİF"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Arial", size=14, bold=True, color="1E3A8A")
    font_subtitle = Font(name="Arial", size=9, italic=True, color="475569")
    font_section = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Arial", size=10, bold=True)
    font_regular = Font(name="Arial", size=10)
    
    fill_section = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    fill_zebra = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    thin_side = Side(border_style="thin", color="CBD5E1")
    double_side = Side(border_style="double", color="475569")
    border_total = Border(top=thin_side, bottom=double_side)
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    
    ws.merge_cells("A1:D1")
    ws["A1"] = "ERC SAMSUN OTOMOTİV SAN. VE TİC. A.Ş."
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A2:D2")
    ws["A2"] = "Eğercili Mah. Atatürk Bulv. No:122/3 Çarşamba / SAMSUN"
    ws["A2"].font = font_regular
    ws["A2"].alignment = Alignment(horizontal="center")
    
    ws.merge_cells("A3:D3")
    ws["A3"] = "Mersis No: 0338084534400001 | Tel: 0362 834 00 55"
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = Alignment(horizontal="center")
    
    ws["A5"] = "PROFORMA NO:"
    ws["A5"].font = font_bold
    ws["B5"] = proforma.invoice_number
    ws["B5"].font = font_regular
    
    ws["C5"] = "TARİH:"
    ws["C5"].font = font_bold
    ws["D5"] = proforma.date.strftime("%d.%m.%Y") if proforma.date else ""
    ws["D5"].font = font_regular
    
    ws["A6"] = "GEÇERLİLİK TARİHİ:"
    ws["A6"].font = font_bold
    ws["B6"] = proforma.validity_date.strftime("%d.%m.%Y") if proforma.validity_date else ""
    ws["B6"].font = font_regular
    
    ws.merge_cells("A8:D8")
    ws["A8"] = "MÜŞTERİ BİLGİLERİ"
    ws["A8"].font = font_section
    ws["A8"].fill = fill_section
    ws["A8"].alignment = Alignment(horizontal="left", indent=1)
    
    ws["A9"] = "Müşteri Ünvanı:"
    ws["A9"].font = font_bold
    ws.merge_cells("B9:D9")
    ws["B9"] = customer.company_name
    ws["B9"].font = font_regular
    
    ws["A10"] = "Vergi Dairesi:"
    ws["A10"].font = font_bold
    ws["B10"] = customer.vergi_dairesi
    ws["B10"].font = font_regular
    
    ws["C10"] = "Vergi Numarası:"
    ws["C10"].font = font_bold
    ws["D10"] = customer.tax_number
    ws["D10"].font = font_regular
    
    ws["A11"] = "Adres:"
    ws["A11"].font = font_bold
    ws.merge_cells("B11:D11")
    ws["B11"] = f"{customer.address or ''} {customer.district or ''} / {customer.city or ''}"
    ws["B11"].font = font_regular
    
    ws.merge_cells("A13:D13")
    ws["A13"] = "ARAÇ TEKNİK BİLGİLERİ"
    ws["A13"].font = font_section
    ws["A13"].fill = fill_section
    ws["A13"].alignment = Alignment(horizontal="left", indent=1)
    
    ws["A14"] = "Araç Modeli:"
    ws["A14"].font = font_bold
    ws.merge_cells("B14:D14")
    ws["B14"] = proforma.vehicle_model
    ws["B14"].font = font_regular
    
    ws["A15"] = "Model Yılı:"
    ws["A15"].font = font_bold
    ws["B15"] = proforma.model_year or ""
    ws["B15"].font = font_regular
    
    ws["C15"] = "Renk:"
    ws["C15"].font = font_bold
    ws["D15"] = proforma.color or ""
    ws["D15"].font = font_regular
    
    ws["A16"] = "Şasi Numarası:"
    ws["A16"].font = font_bold
    ws["B16"] = proforma.chassis_no or ""
    ws["B16"].font = font_regular
    
    ws["C16"] = "Motor Numarası:"
    ws["C16"].font = font_bold
    ws["D16"] = proforma.motor_no or ""
    ws["D16"].font = font_regular
    
    ws["A17"] = "Motor Gücü:"
    ws["A17"].font = font_bold
    ws["B17"] = proforma.motor_power or ""
    ws["B17"].font = font_regular
    
    ws["C17"] = "Azami Ağırlık:"
    ws["C17"].font = font_bold
    ws["D17"] = proforma.max_weight or ""
    ws["D17"].font = font_regular
    
    ws.merge_cells("A19:D19")
    ws["A19"] = "FİNANSAL DETAYLAR VE VERGİLER"
    ws["A19"].font = font_section
    ws["A19"].fill = fill_section
    ws["A19"].alignment = Alignment(horizontal="left", indent=1)
    
    ws["A20"] = "Açıklama"
    ws["A20"].font = font_bold
    ws.merge_cells("B20:C20")
    ws["B20"] = "Oran / Tutar Detayı"
    ws["B20"].font = font_bold
    ws["D20"] = "Tutar (TL)"
    ws["D20"].font = font_bold
    ws["D20"].alignment = Alignment(horizontal="right")
    
    ws["A21"] = "Araç Net Matrahı"
    ws["A21"].font = font_regular
    ws.merge_cells("B21:C21")
    ws["B21"] = "Birim Fiyat"
    ws["B21"].font = font_regular
    ws["D21"] = proforma.unit_price
    ws["D21"].font = font_regular
    ws["D21"].number_format = '#,##0.00'
    ws["D21"].alignment = Alignment(horizontal="right")
    
    ws["A22"] = "Özel Tüketim Vergisi (ÖTV)"
    ws["A22"].font = font_regular
    ws.merge_cells("B22:C22")
    ws["B22"] = f"% {proforma.otv_rate}"
    ws["B22"].font = font_regular
    ws["D22"] = proforma.otv_amount
    ws["D22"].font = font_regular
    ws["D22"].number_format = '#,##0.00'
    ws["D22"].alignment = Alignment(horizontal="right")
    
    ws["A23"] = "ÖTV'li Ara Toplam"
    ws["A23"].font = font_bold
    ws.merge_cells("B23:C23")
    ws["B23"] = "Matrah + ÖTV"
    ws["B23"].font = font_regular
    ws["D23"] = proforma.subtotal
    ws["D23"].font = font_bold
    ws["D23"].number_format = '#,##0.00'
    ws["D23"].alignment = Alignment(horizontal="right")
    
    ws["A24"] = "Katma Değer Vergisi (KDV)"
    ws["A24"].font = font_regular
    ws.merge_cells("B24:C24")
    ws["B24"] = f"% {proforma.kdv_rate}"
    ws["B24"].font = font_regular
    ws["D24"] = proforma.kdv_amount
    ws["D24"].font = font_regular
    ws["D24"].number_format = '#,##0.00'
    ws["D24"].alignment = Alignment(horizontal="right")
    
    ws["A25"] = "Anahtar Teslim Toplam Fiyat"
    ws["A25"].font = font_bold
    ws["A25"].fill = fill_zebra
    ws.merge_cells("B25:C25")
    ws["B25"] = "HER ŞEY DAHİL NET"
    ws["B25"].font = font_bold
    ws["B25"].fill = fill_zebra
    ws["D25"] = proforma.grand_total
    ws["D25"].font = font_bold
    ws["D25"].fill = fill_zebra
    ws["D25"].number_format = '#,##0.00'
    ws["D25"].alignment = Alignment(horizontal="right")
    ws["D25"].border = border_total
    
    ws.merge_cells("A26:D26")
    ws["A26"] = f"Yalnız: {proforma.grand_total_words}"
    ws["A26"].font = font_bold
    ws["A26"].alignment = Alignment(horizontal="left", indent=1)
    
    ws.merge_cells("A28:D28")
    ws["A28"] = "SATIŞ KOŞULLARI VE AÇIKLAMALAR"
    ws["A28"].font = font_section
    ws["A28"].fill = fill_section
    ws["A28"].alignment = Alignment(horizontal="left", indent=1)
    
    ws["A29"] = "Teslim Yeri:"
    ws["A29"].font = font_bold
    ws.merge_cells("B29:D29")
    ws["B29"] = proforma.delivery_place
    ws["B29"].font = font_regular
    
    ws["A30"] = "Ödeme Şekli:"
    ws["A30"].font = font_bold
    ws.merge_cells("B30:D30")
    ws["B30"] = proforma.payment_terms
    ws["B30"].font = font_regular
    
    ws["A31"] = "Notlar & Açıklama:"
    ws["A31"].font = font_bold
    ws.merge_cells("B31:D33")
    ws["B31"] = proforma.notes
    ws["B31"].font = font_regular
    ws["B31"].alignment = Alignment(wrap_text=True, vertical="top")
    
    for r in range(5, 34):
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c)
            if r in [9,10,11, 14,15,16,17, 21,22,23,24,25, 29,30,31,32,33]:
                cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    filename = f"PROFORMA_{proforma.invoice_number}.xlsx"
    headers_dict = {
        'Content-Disposition': f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_dict
    )


@router.put("/proformas/{proforma_id}", response_model=ProformaResponse)
def update_proforma(
    proforma_id: int,
    data: ProformaUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Proforma faturayı günceller."""
    return CRMService(db).update_proforma(proforma_id, data)


@router.delete("/proformas/{proforma_id}")
def delete_proforma(proforma_id: int, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Proforma faturayı siler."""
    CRMService(db).delete_proforma(proforma_id)
    return {"message": "Proforma fatura başarıyla silindi"}


# ── Vehicle Catalog Endpoints ──────────────────────────────────────────

@router.get("/vehicles", response_model=List[VehicleResponse])
def search_vehicles(query: str = "", db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    """Katalogdaki araçları arar."""
    return CRMService(db).search_vehicles(query)


@router.post("/vehicles/import")
async def import_vehicles(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """Excel veya CSV dosyasından araç kataloğunu içe aktarır."""
    return await import_vehicles_from_file(file, db)


